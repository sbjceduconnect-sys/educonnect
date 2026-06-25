from django.contrib.auth import get_user_model, authenticate
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from core.response import api_success, api_error
from .serializers import RegisterSerializer, LoginSerializer, UserSerializer, UserProfileSerializer
from .models import UserProfile

User = get_user_model()


def _get_tokens(user):
    import uuid
    refresh = RefreshToken.for_user(user)
    if user.role == 'student':
        session_id = uuid.uuid4().hex
        user.current_session_id = session_id
        user.save(update_fields=['current_session_id', 'updated_at'])
        refresh['current_session_id'] = session_id
        refresh.access_token['current_session_id'] = session_id
    return str(refresh.access_token), str(refresh)


class RegisterView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if not serializer.is_valid():
            return api_error("Registration failed", errors=serializer.errors, status=400)
        user = serializer.save()
        return api_success(
            data=UserSerializer(user).data,
            message="Registration successful. Awaiting admin approval." if not user.is_approved else "Registration successful.",
            status=201,
        )


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email', '').lower().strip()
        password = request.data.get('password', '')

        if not email or not password:
            return api_error("Email and password are required.", status=400)

        user = authenticate(request, username=email, password=password)
        if user is None:
            return api_error("Invalid credentials.", status=401)

        if not user.is_active:
            return api_error("Account is disabled.", status=403)

        if not user.is_approved:
            return api_error("Your account is pending admin approval.", status=403)

        from apps.audit_logs.models import AuditLog
        AuditLog.log(user=user, action="User Login", target="Login successful", request=request)

        access, refresh = _get_tokens(user)
        return api_success(
            data={
                "accessToken":  access,
                "refreshToken": refresh,
                "user": UserSerializer(user).data,
            },
            message="Login successful.",
        )


class RefreshTokenView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        # Read from request body or cookies
        refresh_token = (
            request.data.get('refresh') or 
            request.data.get('refreshToken') or 
            request.COOKIES.get('refreshToken') or 
            request.COOKIES.get('refresh')
        )
        if not refresh_token:
            return api_error("Refresh token is required.", status=400)
        try:
            refresh = RefreshToken(refresh_token)
            user_id = refresh.payload.get('user_id')
            if user_id:
                user = User.objects.get(id=user_id)
                if user.role == 'student':
                    token_session_id = refresh.payload.get('current_session_id')
                    if not token_session_id or token_session_id != user.current_session_id:
                        return api_error("Your session has expired or you have logged in from another device.", status=401)
            
            access_token = refresh.access_token
            token_session_id = refresh.payload.get('current_session_id')
            if token_session_id:
                access_token['current_session_id'] = token_session_id

            return api_success(data={"accessToken": str(access_token)}, message="Token refreshed.")
        except Exception:
            return api_error("Invalid or expired refresh token.", status=401)


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        if user and user.role == 'student':
            user.current_session_id = None
            user.save(update_fields=['current_session_id', 'updated_at'])
        return api_success(message="Logout successful.")


class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request):
        old_password = request.data.get('old_password') or request.data.get('oldPassword')
        new_password = request.data.get('new_password') or request.data.get('newPassword')
        if not old_password or not new_password:
            return api_error("Both old and new passwords are required.", status=400)
        
        user = request.user
        if not user.check_password(old_password):
            return api_error("Invalid old password.", status=400)
            
        user.set_password(new_password)
        user.save()
        return api_success(message="Password changed successfully.")


class ResetPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, token=None):
        from django.utils import timezone
        
        email = request.data.get('email', '').strip().lower()
        code = request.data.get('code', '').strip()
        password = request.data.get('password')
        
        if not email or not code or not password:
            return api_error("Email, verification code, and password are required.", status=400)
            
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return api_error("Invalid email, verification code, or code has expired.", status=400)
            
        if not user.reset_code or user.reset_code != code:
            return api_error("Invalid email, verification code, or code has expired.", status=400)
            
        if not user.reset_code_expires_at or timezone.now() > user.reset_code_expires_at:
            return api_error("Invalid email, verification code, or code has expired.", status=400)
            
        user.set_password(password)
        user.reset_code = None
        user.reset_code_expires_at = None
        user.save(update_fields=['password', 'reset_code', 'reset_code_expires_at'])
        
        return api_success(message="Password reset successfully.")


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return api_success(data=UserSerializer(request.user).data)


class ForgotPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        import random
        import datetime
        from django.utils import timezone
        
        email = request.data.get('email', '').strip().lower()
        if not email:
            return api_error("Email is required.", status=400)
            
        try:
            user = User.objects.get(email=email)
            code = f"{random.randint(100000, 999999)}"
            user.reset_code = code
            user.reset_code_expires_at = timezone.now() + datetime.timedelta(minutes=10)
            user.save(update_fields=['reset_code', 'reset_code_expires_at'])
            
            print("\n" + "="*60)
            print(f"[PASSWORD RESET] Verification code for {email} is: {code}")
            print("="*60 + "\n")
            
            # Send the real email containing the code if SMTP is configured
            from django.core.mail import send_mail
            from django.conf import settings
            
            subject = "EduConnect - Password Reset Verification Code"
            message = f"Hello,\n\nYour password reset verification code is: {code}\n\nThis code will expire in 10 minutes.\n\nBest regards,\nEduConnect Team"
            
            try:
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
            except Exception as e:
                print(f"Failed to send email: {str(e)}")
            
            return api_success(
                message="Verification code sent to your email."
            )
        except User.DoesNotExist:
            return api_success(message="If this email exists, a verification code will be sent.")


class UserListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        role = request.query_params.get('role')
        is_approved = request.query_params.get('is_approved') or request.query_params.get('isApproved')

        from core.permissions import IsAdmin
        if not IsAdmin().has_permission(request, self):
            if role != 'teacher':
                return api_success(data=UserSerializer(request.user).data)

        qs = User.objects.all()
        if role:
            qs = qs.filter(role=role)
        if is_approved is not None:
            qs = qs.filter(is_approved=(is_approved.lower() == 'true'))
            
        return api_success(
            data=UserSerializer(qs, many=True).data,
        )


class UserDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_user(self, pk):
        try:
            return User.objects.get(pk=pk)
        except User.DoesNotExist:
            return None

    def get(self, request, pk):
        user = self._get_user(pk)
        if not user:
            return api_error("User not found.", status=404)
        return api_success(data=UserSerializer(user).data)

    def put(self, request, pk):
        user = self._get_user(pk)
        if not user:
            return api_error("User not found.", status=404)
        serializer = UserSerializer(user, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return api_success(data=serializer.data, message="User updated.")
        return api_error("Update failed.", errors=serializer.errors)

    def patch(self, request, pk):
        return self.put(request, pk)

    def delete(self, request, pk):
        from core.permissions import IsAdmin
        if not IsAdmin().has_permission(request, self):
            return api_error("Permission denied.", status=403)
        user = self._get_user(pk)
        if not user:
            return api_error("User not found.", status=404)
        user.delete()
        return api_success(message="User deleted.", status=200)


class ApproveUserView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        from core.permissions import IsAdmin
        if not IsAdmin().has_permission(request, self):
            return api_error("Permission denied.", status=403)
        try:
            user = User.objects.get(pk=pk)
            user.is_approved = True
            user.save()
            return api_success(data=UserSerializer(user).data, message="User approved.")
        except User.DoesNotExist:
            return api_error("User not found.", status=404)


class ChangeRoleView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        from core.permissions import IsAdmin
        if not IsAdmin().has_permission(request, self):
            return api_error("Permission denied.", status=403)
        role = request.data.get('role')
        if not role or role not in ['admin', 'teacher', 'student']:
            return api_error("Invalid role.", status=400)
        try:
            user = User.objects.get(pk=pk)
            user.role = role
            user.save()
            return api_success(data=UserSerializer(user).data, message="Role updated successfully.")
        except User.DoesNotExist:
            return api_error("User not found.", status=404)


class BulkApproveView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from core.permissions import IsAdmin
        if not IsAdmin().has_permission(request, self):
            return api_error("Permission denied.", status=403)
        ids = request.data.get('ids', [])
        count = User.objects.filter(id__in=ids).update(is_approved=True)
        return api_success(data={"approved": count}, message=f"{count} users approved.")


class BulkImportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from core.permissions import IsAdmin
        if not IsAdmin().has_permission(request, self):
            return api_error("Permission denied.", status=403)
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return api_error("No file uploaded.", status=400)
        
        imported_count = 0
        try:
            is_excel = file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls')
            rows = []
            if is_excel:
                import openpyxl
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                for row_idx in range(2, sheet.max_row + 1):
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        if header:
                            row_data[header] = sheet.cell(row=row_idx, column=col_idx + 1).value
                    if any(row_data.values()):
                        rows.append(row_data)
            else:
                import csv
                import io
                data = file_obj.read().decode('utf-8-sig')
                io_string = io.StringIO(data)
                reader = csv.DictReader(io_string)
                for row in reader:
                    rows.append(row)

            for row in rows:
                username = row.get('username')
                email = row.get('email')
                first_name = row.get('first_name') or row.get('firstName', '')
                last_name = row.get('last_name') or row.get('lastName', '')
                role = row.get('role', 'student')
                password = row.get('password', 'Edu12345')
                
                if not username or not email:
                    continue
                
                if User.objects.filter(email=email).exists() or User.objects.filter(username=username).exists():
                    continue
                
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    role=role,
                    is_approved=True
                )
                user.set_password(password)
                user.save()
                
                # Setup profile
                profile, _ = UserProfile.objects.get_or_create(user=user)
                if role == 'student':
                    profile.enrollment_no = row.get('enrollment_no') or row.get('enrollmentNo', '')
                    profile.stream = row.get('stream', 'Science')
                    profile.section = row.get('section', 'A')
                else:
                    profile.employee_id = row.get('employee_id') or row.get('employeeId', '')
                    profile.qualification = row.get('qualification', '')
                    profile.specialization = row.get('specialization', '')
                profile.save()
                imported_count += 1
            
            return api_success(data={"imported": imported_count}, message=f"Successfully imported {imported_count} users.", status=201)
        except ImportError:
            return api_error("Excel parser is not available. Please upload a CSV file.", status=400)
        except Exception as e:
            return api_error(f"Failed to process file: {str(e)}", status=400)


class AdminDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from core.permissions import IsAdmin
        if not IsAdmin().has_permission(request, self):
            return api_error("Permission denied.", status=403)
        
        from apps.courses.models import Course
        
        total_students = User.objects.filter(role='student').count()
        total_teachers = User.objects.filter(role='teacher').count()
        total_admins = User.objects.filter(role='admin').count()
        pending_approvals = User.objects.filter(is_approved=False).count()
        total_courses = Course.objects.count()
        total_enrolled = User.objects.filter(enrolled_courses__isnull=False).distinct().count()
        
        recent_users = User.objects.all().order_by('-created_at')[:10]
        
        stats = {
            "totalStudents": total_students,
            "totalTeachers": total_teachers,
            "totalAdmins": total_admins,
            "pendingApprovals": pending_approvals,
            "totalCourses": total_courses,
            "totalEnrolledStudents": total_enrolled,
        }
        
        return api_success(data={
            "stats": stats,
            "recentUsers": UserSerializer(recent_users, many=True).data
        })


class TeacherDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.courses.models import Course
        from apps.announcements.models import Announcement
        
        teacher = request.user
        courses = Course.objects.filter(teacher=teacher)
        total_students = User.objects.filter(enrolled_courses__in=courses).distinct().count()
        announcements = Announcement.objects.all()
        
        serialized_courses = []
        for course in courses:
            serialized_courses.append({
                "id": course.id,
                "title": course.name,
                "courseCode": course.code,
                "enrolledStudentIds": list(course.students.values_list('id', flat=True))
            })
            
        return api_success(data={
            "totalCourses": courses.count(),
            "totalStudents": total_students,
            "courses": serialized_courses,
            "announcements": list(announcements.values('id', 'title', 'created_at'))
        })


class StudentDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.courses.models import Course
        from apps.announcements.models import Announcement
        from apps.attendance.models import AttendanceRecord
        from apps.exams.models import Result
        
        student = request.user
        enrolled_courses = Course.objects.filter(students=student)
        
        # Attendance stats
        attendance_qs = AttendanceRecord.objects.filter(student=student)
        total_classes = attendance_qs.count()
        attended = attendance_qs.filter(status__in=['Present', 'Late']).count()
        attendance_percentage = round((attended / total_classes * 100), 1) if total_classes else 0
        
        attendance_stats = {
            "percentage": attendance_percentage,
            "attended": attended,
            "totalClasses": total_classes
        }
        
        # Performance/Results
        recent_results = []
        performance_stats = {
            "passed": 0,
            "totalSubjects": 0,
            "averagePercentage": 0,
            "overallGrade": "Locked"
        }
        
        if student.fees_paid:
            results = Result.objects.filter(student=student)
            passed_count = 0
            total_pct = 0
            
            for r in results:
                pct = (r.marks / r.exam.max_marks * 100) if r.exam.max_marks else 0
                if pct >= 40:
                    passed_count += 1
                total_pct += pct
                recent_results.append({
                    "id": r.id,
                    "examTitle": r.exam.title,
                    "marksObtained": r.marks,
                    "maxMarks": r.exam.max_marks,
                    "percentage": round(pct, 1),
                    "grade": r.grade
                })
                
            average_pct = round(total_pct / results.count(), 1) if results.count() else 0
            overall_grade = 'A+' if average_pct>=90 else 'A' if average_pct>=80 else 'B+' if average_pct>=75 else 'B' if average_pct>=70 else 'C' if average_pct>=60 else 'D' if average_pct>=50 else 'F' if results.exists() else 'N/A'
            
            performance_stats = {
                "passed": passed_count,
                "totalSubjects": results.count(),
                "averagePercentage": average_pct,
                "overallGrade": overall_grade
            }
        
        # Announcements
        announcements = Announcement.objects.all().order_by('-created_at')[:5]
        serialized_announcements = []
        for ann in announcements:
            serialized_announcements.append({
                "id": ann.id,
                "title": ann.title,
                "createdAt": ann.created_at,
                "priority": getattr(ann, 'priority', 'normal')
            })
            
        # Courses
        serialized_courses = []
        for course in enrolled_courses:
            serialized_courses.append({
                "id": course.id,
                "title": course.name,
                "courseCode": course.code
            })
            
        return api_success(data={
            "feesPaid": student.fees_paid,
            "enrolledCourses": enrolled_courses.count(),
            "attendance": attendance_stats,
            "performance": performance_stats,
            "recentResults": recent_results,
            "courses": serialized_courses,
            "announcements": serialized_announcements
        })


class BulkDeleteUsersView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from core.permissions import IsAdmin
        from django.conf import settings
        if not IsAdmin().has_permission(request, self):
            return api_error("Permission denied.", status=403)

        class_filter = request.data.get('class_filter') or request.data.get('classFilter')
        security_code = request.data.get('security_code') or request.data.get('securityCode')

        expected_code = getattr(settings, 'BULK_DELETE_SECURITY_CODE', 'SBJC-DELETE-CONFIRM-2026')
        if security_code != expected_code:
            return api_error("Invalid security code.", status=400)

        if class_filter not in ['11th', '12th', 'all']:
            return api_error("Invalid class filter. Must be '11th', '12th', or 'all'.", status=400)

        # Only delete users with role='student'
        queryset = User.objects.filter(role='student')

        if class_filter == '11th':
            queryset = queryset.filter(profile__stream__startswith='11th')
        elif class_filter == '12th':
            queryset = queryset.filter(profile__stream__startswith='12th')

        count = queryset.count()
        queryset.delete()

        return api_success(data={"deleted": count}, message=f"Successfully bulk deleted {count} students.")

